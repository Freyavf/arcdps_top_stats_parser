#!/usr/bin/env python3

from stat_classes import *
import xlrd
from xlutils.copy import copy
import jsons
import json
import pandas as pd
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
    )
from openpyxl.utils import get_column_letter


# get the professions of all players indicated by the indices. Additionally, get the length of the longest profession name.
# Input:
# players = list of all players
# indices = list of relevant indices
# config = config to use for top stats computation/printing
# Output:
# list of profession strings, maximum profession length
def get_professions_and_length(players, indices, config):
    profession_strings = list()
    profession_length = 0
    for i in indices:
        player = players[i]
        professions_str = config.profession_abbreviations[player.profession]
        profession_strings.append(professions_str)
        if len(professions_str) > profession_length:
            profession_length = len(professions_str)
    return profession_strings, profession_length



# get total duration in h, m, s
def get_total_fight_duration_in_hms(fight_duration_in_s):
    total_fight_duration = {}
    total_fight_duration['h'] = int(fight_duration_in_s/3600)
    total_fight_duration['m'] = int((fight_duration_in_s - total_fight_duration['h']*3600) / 60)
    total_fight_duration['s'] = int(fight_duration_in_s - total_fight_duration['h']*3600 -  total_fight_duration['m']*60)
    return total_fight_duration



# prints output_string to the console and the output_file, with a linebreak at the end
def myprint(output_file, output_string, log_level, config = None):
    if config != None:
        if log_level == "warning" and config.log_level == "info":
            return
        if log_level == "debug" and config.log_level != "debug":
            return
        print(output_string)
        output_file.write(output_string+"\n")
        
    if config == None:
        print(output_string)
        output_file.write(output_string+"\n")


# checks whether the given column contains string values
def is_string_column(column_name):
    if column_name in ["account", "name", "profession"]:
        return True
    else:
        return False

    
# Write the top x people who achieved top total stat.
# Input:
# players = list of Players
# top_players = list of indices in players that are considered as top
# stat = which stat are we considering
# xls_output_filename = where to write to
def write_stats_xls(players, top_players, stat, xls_output_filename, config):
    writer = pd.ExcelWriter(xls_output_filename, engine = "openpyxl", mode = 'a')

    sorting_columns = config.sort_xls_by[stat]

    # sort in descending order, unless it's a stat where low values are good and total or avg are sorted
    sort_ascending = [False for x in config.sort_xls_by[stat]]
    if stat == 'deaths' or stat == 'stripped' or stat == 'dist' or 'dmg_taken' in stat:
        for i, val in enumerate(config.sort_xls_by[stat]):
            if val == "avg" or val == "total":
                sort_ascending[i] = True
    # always sort strings in ascending order
    for i, val in enumerate(config.sort_xls_by[stat]):
        if is_string_column(val):
            sort_ascending[i] = True
                
    df = create_panda_dataframe(players, top_players, stat, sorting_columns, sort_ascending, config)
    
    df.to_excel(writer, sheet_name = config.stat_names[stat], startrow = 3, index = False, header = False)
    book = writer.book
    sheet = book[config.stat_names[stat]]
    bold = Font(bold=True)
    bold_red = Font(bold=True, color='FF0000')
    bold_green = Font(bold=True, color='006400')
    sheet['A1'] = config.stat_descriptions[stat]
    sheet['A1'].font = bold

    column_names = [config.xls_column_names[c] for c in list(df) if c in config.xls_column_names]
    column_names.append("Times Top "+str(config.num_players_considered_top[stat]))
    column_names.append("Percentage Top"+str(config.num_players_considered_top[stat]))

    # rename the columns for the xls
    if stat == 'spike_dmg':
        column_names.append("Maximum "+stat)
    else:
        column_names.append("Total "+stat)

    if stat == 'deaths' or stat == 'kills' or stat == 'downs':
        column_names.append("Average "+stat+" per min "+config.duration_for_averages[stat])
    elif stat == 'spike_dmg':
        column_names.append("Average "+stat+" over all fights")
    elif stat in config.squad_buff_ids and stat in config.buffs_not_stacking:
        column_names.append("Average "+stat+" in %")
    elif stat not in config.self_buff_ids:
        column_names.append("Average "+stat+" per s "+config.duration_for_averages[stat])
    for i in range(len(column_names)):
        header_cell = sheet.cell(row=3, column=(i+1))
        header_cell.value = column_names[i]
        header_cell.font = bold

    # make relevant classes bold
    (max_row, max_col) = df.shape
    top_value_per_profession = {profession: -1 for profession in config.relevant_classes[stat]}
    i = 0

    # the actual stat value that is used first in sorting
    stat_sorting_column = 0
    while stat_sorting_column < len(sorting_columns) and is_string_column(sorting_columns[stat_sorting_column]):
        stat_sorting_column = stat_sorting_column+1
    if stat_sorting_column >= len(sorting_columns):
        stat_sorting_column = -1
         
    for _, row in df.iterrows():
        prof = row["profession"]
        # mark all relevant classes in bold
        if prof in config.relevant_classes[stat]:
            for j in range(1,10):
                sheet.cell(i+4, j).font = bold
        i = i + 1

    filters = sheet.auto_filter
    filters.ref = "A3:" + get_column_letter(sheet.max_column) + str(sheet.max_row)

    book.save(xls_output_filename)


    
# Write xls fight overview
# Input:
# fights = list of Fights as returned by collect_stat_data
# overall_squad_stats = overall stats of the whole squad; output of get_overall_squad_stats
# overall_raid_stats = raid stats like start time, end time, total kills, etc.; output of get_overall_raid_stats
# config = the config to use for stats computation
# xls_output_filename = where to write to
def write_fights_overview_xls(fights, overall_squad_stats, overall_raid_stats, config, xls_output_filename):
    writer = pd.ExcelWriter(xls_output_filename, engine = "openpyxl")

    df = create_panda_dataframe_overview(fights, overall_squad_stats, overall_raid_stats, config)
#    print(df)
    df.to_excel(writer, sheet_name = "Fights Overview", index = False)
    book = writer.book
    book.save(xls_output_filename)



# write all stats to a json file
# Input:
# overall_raid_stats = raid stats like start time, end time, total kills, etc.; output of get_overall_raid_stats
# overall_squad_stats = overall stats of the whole squad; output of get_overall_squad_stats
# fights = list of Fights
# config = the config used for stats computation
# output = file to write to

def write_to_json(overall_raid_stats, overall_squad_stats, fights, players, top_total_stat_players, top_average_stat_players, top_consistent_stat_players, top_percentage_stat_players, output_file):
    json_dict = {}
    json_dict["overall_raid_stats"] = {key: value for key, value in overall_raid_stats.items()}
    json_dict["overall_squad_stats"] = {key: value for key, value in overall_squad_stats.items()}
    json_dict["fights"] = [jsons.dump(fight) for fight in fights]
    json_dict["players"] = [jsons.dump(player) for player in players]
    json_dict["top_total_players"] =  {key: value for key, value in top_total_stat_players.items()}
    json_dict["top_average_players"] =  {key: value for key, value in top_average_stat_players.items()}
    json_dict["top_consistent_players"] =  {key: value for key, value in top_consistent_stat_players.items()}
    json_dict["top_percentage_players"] =  {key: value for key, value in top_percentage_stat_players.items()}

    with open(output_file, 'w') as json_file:
        json.dump(json_dict, json_file, indent=4)



# Create a panda dataframe for a fights overview
def create_panda_dataframe_overview(fights, overall_squad_stats, overall_raid_stats, config):
    first_col = ["" for i in range(len(fights))]
    fight_num = [i for i in range(len(fights))]
    start_date = [fight.start_time.split()[0] for fight in fights]
    start_time = [fight.start_time.split()[1] for fight in fights]
    end_time = [fight.end_time.split()[1] for fight in fights]
    duration = [fight.duration for fight in fights]
    skipped = ["yes" if fight.skipped else "no" for fight in fights]
    num_allies = [fight.allies for fight in fights]
    num_enemies = [fight.enemies for fight in fights]
    kills = [fight.kills for fight in fights]
    stats = {}
    # for squad buffs and distance, total values don't make sense
    for stat in config.stats_to_compute:
        if stat not in config.squad_buff_ids and stat != "dist":
            stats[stat] = [fight.total_stats[stat] for fight in fights]
        else:
            stats[stat] = [fight.avg_stats[stat] for fight in fights]
    first_col.append("Sum/Avg. in used fights")
    fight_num.append(overall_raid_stats['num_used_fights'])
    start_date.append(overall_raid_stats['date'])
    start_time.append(overall_raid_stats['start_time'])
    end_time.append(overall_raid_stats['end_time'])    
    duration.append(overall_raid_stats['used_fights_duration'])
    skipped.append(overall_raid_stats['num_skipped_fights'])
    num_allies.append(overall_raid_stats['mean_allies'])    
    num_enemies.append(overall_raid_stats['mean_enemies'])
    kills.append(overall_raid_stats['total_kills'])
    for stat in config.stats_to_compute:
        stats[stat].append(overall_squad_stats['total'][stat])

    data = {"": first_col,
            "#": fight_num,
            "Date": start_date,
            "Start Time": start_time,
            "End Time": end_time,
            "Duration in s": duration,
            "Skipped": skipped,
            "Num. Allies": num_allies,
            "Num. Enemies": num_enemies,
            "Kills": kills
            }
    for stat in stats:
        data[config.stat_names[stat]] = stats[stat]
        
    df = pd.DataFrame(data)
    return df
        
    
# Create a panda dataframe from stat data
# Input:
# players = list of Players
# top_players = list of indices in players that are considered as top
# stat = which stat are we considering
# sorting_column = which column to sort by
# sort_ascending = are we sorting in ascending order
# config = config of how the stats are computed
# Output:
# panda data frame containing data to be written to an excel sheet
def create_panda_dataframe(players, top_players, stat, sorting_columns, sort_ascending, config):
    accounts = (players[top_players[i]].account for i in range(len(top_players)))
    names = (players[top_players[i]].name for i in range(len(top_players)))
    professions = (players[top_players[i]].profession for i in range(len(top_players)))
    num_fights_present = (players[top_players[i]].num_fights_present for i in range(len(top_players)))
    duration_present = (players[top_players[i]].duration_present[config.duration_for_averages[stat]] for i in range(len(top_players)))
    consistency_stats = (players[top_players[i]].consistency_stats[stat] for i in range(len(top_players)))
    portion_top_stats = (players[top_players[i]].portion_top_stats[stat]*100 for i in range(len(top_players)))
    total_stats = (players[top_players[i]].total_stats[stat] for i in range(len(top_players)))
    average_stats = list()
    if stat not in config.self_buff_ids:
        average_stats = (players[top_players[i]].average_stats[stat] for i in range(len(top_players)))
    data = {"account": accounts,
            "name": names,
            "profession": professions,
            "attendance_num": num_fights_present,
            "attendance_duration": duration_present,
            "times_top": consistency_stats,
            "percentage_top": portion_top_stats,
            "total": total_stats}
    if stat not in config.self_buff_ids:
        data["avg"] = average_stats
    
    df = pd.DataFrame(data)
    print(stat)
    #if stat == 'interrupts':
    #    print(df)

    df.sort_values(sorting_columns, ascending=sort_ascending, inplace=True)
    return df
